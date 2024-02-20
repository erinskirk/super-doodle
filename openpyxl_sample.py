import openpyxl 

dir_path = 'C:/Users/esk7/OneDrive - PGE/Desktop/GPOM QA Dispatch_02192024.xlsx'
wb = openpyxl.load_workbook(dir_path)
ws = wb.active

print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

print('The value in cell A1 is: '+ws['A1'].value)

values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)

data=[ws.cell(row=i,column=2).value for i in range(2,12)]
print(data)

#create list
my_list = list()

#write the first ten rows in a range of columns (6 in this case) in the spreadsheet
for value in ws.iter_rows(
    min_row=1, max_row=11, min_col=1, max_col=6, 
    values_only=True):
    my_list.append(value)
    
for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
    #the < means left align and the numbers in the bracket give us size of column
    (print ("{:<30}{:<35}{:<35}{:<25}{:<20}{:<35}".format(ele1,ele2,ele3,ele4,ele5,ele6)))
input('Confirm by pressing any key')
